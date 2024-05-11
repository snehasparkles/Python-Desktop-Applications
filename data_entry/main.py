from tkinter import *
from tkinter.ttk import Combobox
from tkinter import messagebox
import openpyxl,xlrd
from openpyxl import Workbook
import pathlib
import tkinter as tk

PINK = "#e2979c"
RED = "#e7305b"
GREEN = "#9bdeac"
YELLOW = "#f7f5dd"
FONT_NAME = "Courier"



screen = Tk()
screen.title("DATA ENTRY")
screen.config(bg=GREEN)
# screen.minsize(width=1000,height=400)
screen.geometry('700x400+300+200')
screen.resizable(False,False)


file = pathlib.Path("backend_data.xlsx")
if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet['A1'] = "FULL NAME"
    sheet['B1'] = "PHONE NUMBER"
    sheet['C1'] = "AGE"
    sheet['D1'] = "GENDER"
    sheet['E1'] = "ADDRESS"
    file.save("backend_data.xlsx")


#submit
def submit():
    name = namevalue.get()
    contact = convalue.get()
    age = agevalue.get()
    gender = gender_input.get()
    address = address_entry.get(1.0,END)
    
    if name == "" or contact == "" or age == "" or gender == "" or address == "":
        messagebox.showinfo(title="Oops",message="Kindly fill all the data")
    else:
 
        file = openpyxl.load_workbook("backend_data.xlsx")
        sheet = file.active
        sheet.cell(column = 1,row = sheet.max_row + 1,value = name)
        sheet.cell(column = 2,row = sheet.max_row ,value = contact)
        sheet.cell(column = 3,row = sheet.max_row ,value = age)
        sheet.cell(column = 4,row = sheet.max_row ,value = gender)
        sheet.cell(column = 5,row = sheet.max_row ,value = address)
        
        file.save(r"backend_data.xlsx")
        
        messagebox.showinfo(title="Info",message="Details added successfully!")
        clear()




#clear
def clear():
    namevalue.set("")
    convalue.set("")
    agevalue.set("")
    address_entry.delete(1.0,END)
    



#icon
icon = PhotoImage(file="computer-user-icon-28.png")
screen.iconphoto(False,icon)

msg = Label(text="please fill out the entry form",font=("arial",14),bg=GREEN)
msg.place(x=20,y=20)

name = Label(text="Name",font=("arial",14),bg=GREEN)
name.place(x=50,y=100)

namevalue = tk.StringVar()
convalue =  tk.StringVar() 
agevalue =  tk.StringVar()


name_input = Entry(screen,textvariable=namevalue,width=45,bd=2,font=20)
# name_input.insert(END,string="Enter your Name")
# print(name_input.get())
name_input.place(x=200,y=100)

contact = Label(text="Contact No",font=("arial",14),bg=GREEN)
contact.place(x=50,y=150)

contact_input = Entry(screen,textvariable=convalue,width=45,bd=2,font=20)
# contact_input.insert(END,string="Enter mobile no")
# print(name_input.get())
contact_input.place(x=200,y=150)


age = Label(text="Age",font=("arial",14),bg=GREEN)
age.place(x=50,y=200)

age_input = Entry(screen,textvariable=agevalue,width=15,bd=2,font=20)
# age_input.insert(END,string="age")
# print(name_input.get())
age_input.place(x=200,y=200)

gender = Label(text="Gender",font=("arial",14),bg=GREEN)
gender.place(x=370,y=200)

gender_input = Combobox(screen,values=['Male','Female'],font='arial 14',state='r',width=14)
gender_input.place(x=440,y=200)
gender_input.set('Male')

address = Label(text="Address",font=("arial",14),bg=GREEN)
address.place(x=50,y=250)

address_entry = Text(screen,width=50,height=4,bd=2)
address_entry.place(x=200,y=250)

subbtn = Button(text="Submit",bg=GREEN,width=15,height=1,command=submit)
subbtn.place(x=200,y=350)

clrbtn = Button(text="Clear",bg=GREEN,width=15,height=1,command=clear)
clrbtn.place(x=340,y=350)

extbtn = Button(text="Exit",bg=GREEN,width=15,height=1,command= lambda:screen.destroy())
extbtn.place(x=480,y=350)

screen.mainloop()
